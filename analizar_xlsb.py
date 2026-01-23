#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para analizar archivos XLSB y extraer estructura de hojas
"""

import sys
import json
from pathlib import Path

# Intentar importar librerías necesarias
try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from pycel import ExcelCompiler
    PYCEL_AVAILABLE = True
except ImportError:
    PYCEL_AVAILABLE = False

try:
    from openpyxl_image_loader import load_workbook as load_wb_images
    IMAGE_LOADER_AVAILABLE = True
except ImportError:
    IMAGE_LOADER_AVAILABLE = False


def analizar_xlsb_con_pandas(ruta_archivo):
    """Analiza XLSB usando pandas (más flexible para XLSB)"""
    try:
        print(f"\n[PANDAS] Leyendo {ruta_archivo}...")
        
        # Leer todas las hojas
        hojas = pd.read_excel(ruta_archivo, sheet_name=None)
        
        resultado = {
            'archivo': str(ruta_archivo),
            'metodo': 'pandas',
            'hojas': {}
        }
        
        for nombre_hoja, df in hojas.items():
            print(f"\n  Hoja: {nombre_hoja}")
            print(f"    Dimensiones: {df.shape[0]} filas x {df.shape[1]} columnas")
            print(f"    Columnas: {list(df.columns)}")
            
            # Mostrar primeras filas
            if df.shape[0] > 0:
                print(f"    Primeras 3 filas:")
                for idx, row in df.head(3).iterrows():
                    print(f"      {idx}: {row.to_dict()}")
            
            resultado['hojas'][nombre_hoja] = {
                'filas': df.shape[0],
                'columnas': df.shape[1],
                'nombre_columnas': list(df.columns),
                'tipos': {col: str(df[col].dtype) for col in df.columns},
                'datos_sample': df.head(3).to_dict('records')
            }
        
        return resultado
        
    except Exception as e:
        print(f"  Error con pandas: {e}")
        return None


def analizar_xlsb_openpyxl(ruta_archivo):
    """Intenta analizar XLSB con openpyxl (soporta XLSB en versiones recientes)"""
    try:
        print(f"\n[OPENPYXL] Leyendo {ruta_archivo}...")
        
        wb = load_workbook(str(ruta_archivo), data_only=True)
        
        resultado = {
            'archivo': str(ruta_archivo),
            'metodo': 'openpyxl',
            'hojas': {}
        }
        
        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            print(f"\n  Hoja: {nombre_hoja}")
            print(f"    Dimensiones: {ws.dimensions}")
            
            # Extraer encabezados (primera fila)
            encabezados = []
            for cell in ws[1]:
                encabezados.append(cell.value)
            
            print(f"    Columnas: {encabezados}")
            
            # Extraer datos
            datos = []
            for row_idx in range(2, min(5, ws.max_row + 1)):  # Primeras 3 filas de datos
                fila = {}
                for col_idx, encabezado in enumerate(encabezados):
                    cell = ws.cell(row=row_idx, column=col_idx + 1)
                    fila[encabezado] = cell.value
                datos.append(fila)
                print(f"      Fila {row_idx}: {fila}")
            
            resultado['hojas'][nombre_hoja] = {
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'encabezados': encabezados,
                'datos_sample': datos
            }
        
        return resultado
        
    except Exception as e:
        print(f"  Error con openpyxl: {e}")
        return None


def main():
    ruta_workspace = Path(r'C:\hibo-cocina')
    
    # Buscar archivos XLSB
    archivos_xlsb = list(ruta_workspace.glob('*.xlsb'))
    
    if not archivos_xlsb:
        print("❌ No se encontraron archivos XLSB en:", ruta_workspace)
        return
    
    print("=" * 80)
    print("ANÁLISIS DE ARCHIVOS XLSB")
    print("=" * 80)
    
    resultados = {}
    
    for archivo_xlsb in archivos_xlsb:
        print(f"\n\n{'='*80}")
        print(f"ARCHIVO: {archivo_xlsb.name}")
        print(f"{'='*80}")
        
        # Intentar primero con pandas (mejor soporte para XLSB)
        if PANDAS_AVAILABLE:
            resultado = analizar_xlsb_con_pandas(str(archivo_xlsb))
            if resultado:
                resultados[archivo_xlsb.name] = resultado
                continue
        
        # Si no funciona pandas, intentar openpyxl
        if OPENPYXL_AVAILABLE:
            resultado = analizar_xlsb_openpyxl(str(archivo_xlsb))
            if resultado:
                resultados[archivo_xlsb.name] = resultado
                continue
        
        print(f"⚠️  No se pudo leer {archivo_xlsb.name}")
        print(f"    Librerías disponibles:")
        print(f"      - pandas: {PANDAS_AVAILABLE}")
        print(f"      - openpyxl: {OPENPYXL_AVAILABLE}")
    
    # Guardar resultados
    if resultados:
        archivo_json = ruta_workspace / 'analisis_xlsb.json'
        with open(archivo_json, 'w', encoding='utf-8') as f:
            json.dump(resultados, f, indent=2, ensure_ascii=False)
        print(f"\n\n✅ Análisis guardado en: {archivo_json}")


if __name__ == '__main__':
    main()
